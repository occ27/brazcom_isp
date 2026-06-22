import io
import zipfile
from datetime import datetime, date
from typing import List, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session

from app.models.models import (
    Cliente, EmpresaCliente, EmpresaClienteEndereco, 
    Servico, ServicoContratado, Receivable, BankAccount, 
    Ticket, CaixaSessao, CaixaMovimentacao
)
from app.models.network import Router

class BackupService:
    @staticmethod
    def _create_excel_file(title: str, headers: List[str], rows: List[List[Any]]) -> bytes:
        """Auxiliar para gerar o arquivo Excel (.xlsx) na memória."""
        wb = Workbook()
        ws = wb.active
        ws.title = title

        # Estilo do cabeçalho
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        ws.append(headers)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Inserção das linhas de dados
        for row in rows:
            formatted_row = []
            for val in row:
                # Formatando datas e enums para representações textuais amigáveis
                if isinstance(val, (datetime, date)):
                    formatted_row.append(val.strftime("%d/%m/%Y %H:%M:%S") if isinstance(val, datetime) else val.strftime("%d/%m/%Y"))
                elif hasattr(val, "value"):  # Enums
                    formatted_row.append(str(val.value))
                elif val is None:
                    formatted_row.append("")
                else:
                    formatted_row.append(val)
            ws.append(formatted_row)

        # Ajuste automático da largura das colunas
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @classmethod
    def generate_company_backup(cls, db: Session, empresa_id: int) -> bytes:
        """Consulta as tabelas principais e gera um arquivo ZIP contendo as planilhas Excel."""
        zip_buffer = io.BytesIO()

        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            
            # --- 1. Clientes ---
            clientes_db = db.query(Cliente, EmpresaClienteEndereco).join(
                EmpresaCliente, Cliente.id == EmpresaCliente.cliente_id
            ).outerjoin(
                EmpresaClienteEndereco, 
                (EmpresaCliente.id == EmpresaClienteEndereco.empresa_cliente_id) & 
                (EmpresaClienteEndereco.is_principal == True)
            ).filter(
                EmpresaCliente.empresa_id == empresa_id
            ).all()

            clientes_headers = [
                "ID", "Nome / Razão Social", "CPF / CNPJ", "E-mail", "Telefone", 
                "Tipo Pessoa", "Ind. IE Dest", "IE", "Ativo", 
                "Endereço", "Número", "Bairro", "Município", "UF", "CEP", "Data Cadastro"
            ]
            clientes_rows = []
            for c, end in clientes_db:
                clientes_rows.append([
                    c.id,
                    c.nome_razao_social,
                    c.cpf_cnpj or c.idOutros or "",
                    c.email or "",
                    c.telefone or "",
                    c.tipo_pessoa,
                    c.ind_ie_dest,
                    c.inscricao_estadual or "",
                    "Sim" if c.is_active else "Não",
                    end.endereco if end else "",
                    end.numero if end else "",
                    end.bairro if end else "",
                    end.municipio if end else "",
                    end.uf if end else "",
                    end.cep if end else "",
                    c.created_at
                ])
            zip_file.writestr("clientes.xlsx", cls._create_excel_file("Clientes", clientes_headers, clientes_rows))

            # --- 2. Serviços / Planos ---
            servicos_db = db.query(Servico).filter(Servico.empresa_id == empresa_id).all()
            servicos_headers = [
                "ID", "Código", "Descrição", "Tipo", "Classificação (cClass)", "Unidade Medida", 
                "Valor Unitário", "Ativo", "CFOP", "NCM", "Download (Mbps)", "Upload (Mbps)", 
                "Limite de Banda", "Fidelidade (Meses)", "Ciclo Cobrança"
            ]
            servicos_rows = []
            for s in servicos_db:
                servicos_rows.append([
                    s.id,
                    s.codigo,
                    s.descricao,
                    s.tipo,
                    s.cClass,
                    s.unidade_medida,
                    s.valor_unitario,
                    "Sim" if s.is_active else "Não",
                    s.cfop or "",
                    s.ncm or "",
                    s.download_speed or "",
                    s.upload_speed or "",
                    s.max_limit or "",
                    s.fidelity_months or "",
                    s.billing_cycle or ""
                ])
            zip_file.writestr("servicos.xlsx", cls._create_excel_file("Serviços", servicos_headers, servicos_rows))

            # --- 3. Contratos ---
            contratos_db = db.query(ServicoContratado).filter(ServicoContratado.empresa_id == empresa_id).all()
            contratos_headers = [
                "ID", "Nº Contrato", "Cliente", "Plano / Serviço", "Status", "Data Início", "Data Fim", 
                "Dia Emissão", "Dia Vencimento", "Valor Unitário", "Valor Total", "IP Designado", 
                "MAC Address", "Conexão", "Autenticação", "PPPoE Username", "ONU Serial", "OLT", "CTO", "Ativo"
            ]
            contratos_rows = []
            for c in contratos_db:
                olt_name = c.olt.nome if c.olt else (c.olt_nome or "")
                cto_name = c.cto.nome if c.cto else (c.cto_nome or "")
                contratos_rows.append([
                    c.id,
                    c.numero_contrato or "",
                    c.cliente.nome_razao_social if c.cliente else "",
                    c.servico.descricao if c.servico else "",
                    c.status,
                    c.d_contrato_ini,
                    c.d_contrato_fim,
                    c.dia_emissao,
                    c.dia_vencimento or "",
                    c.valor_unitario,
                    c.valor_total or "",
                    c.assigned_ip or "",
                    c.mac_address or "",
                    c.tipo_conexao,
                    c.metodo_autenticacao or "",
                    c.pppoe_username or "",
                    c.onu_serial or "",
                    olt_name,
                    cto_name,
                    "Sim" if c.is_active else "Não"
                ])
            zip_file.writestr("contratos.xlsx", cls._create_excel_file("Contratos", contratos_headers, contratos_rows))

            # --- 4. Cobranças / Recebíveis ---
            receivables_db = db.query(Receivable).filter(Receivable.empresa_id == empresa_id).all()
            receivables_headers = [
                "ID", "Cliente", "ID Contrato", "Tipo", "Data Vencimento", "Valor Cobrado", 
                "Valor Pago", "Status", "Banco", "Carteira", "Nosso Número", "Código Barras", 
                "Linha Digitável", "Data Emissão", "Data Pagamento"
            ]
            receivables_rows = []
            for r in receivables_db:
                receivables_rows.append([
                    r.id,
                    r.cliente.nome_razao_social if r.cliente else "",
                    r.servico_contratado_id or "",
                    r.tipo,
                    r.due_date,
                    r.amount,
                    r.paid_amount or "",
                    r.status,
                    r.bank,
                    r.carteira or "",
                    r.nosso_numero or "",
                    r.codigo_barras or "",
                    r.linha_digitavel or "",
                    r.issue_date,
                    r.paid_at
                ])
            zip_file.writestr("cobrancas.xlsx", cls._create_excel_file("Cobranças", receivables_headers, receivables_rows))

            # --- 5. Contas Bancárias ---
            bank_db = db.query(BankAccount).filter(BankAccount.empresa_id == empresa_id).all()
            bank_headers = [
                "ID", "Banco", "Nome Identificador", "Agência", "Conta", "Titular", 
                "Carteira", "Convênio", "Ativo", "Padrão"
            ]
            bank_rows = []
            for b in bank_db:
                agencia = f"{b.agencia or ''}-{b.agencia_dv or ''}" if b.agencia_dv else (b.agencia or "")
                conta = f"{b.conta or ''}-{b.conta_dv or ''}" if b.conta_dv else (b.conta or "")
                bank_rows.append([
                    b.id,
                    b.bank,
                    b.name or "",
                    agencia,
                    conta,
                    b.titular or "",
                    b.carteira or "",
                    b.convenio or "",
                    "Sim" if b.is_active else "Não",
                    "Sim" if b.is_default else "Não"
                ])
            zip_file.writestr("contas_bancarias.xlsx", cls._create_excel_file("Contas Bancárias", bank_headers, bank_rows))

            # --- 6. Tickets de Suporte ---
            tickets_db = db.query(Ticket).filter(Ticket.empresa_id == empresa_id).all()
            tickets_headers = [
                "ID", "Título", "Cliente", "ID Contrato", "Status", "Prioridade", "Categoria", 
                "Descrição", "Resolução", "Criado Por", "Atribuído Para", "Criado Em", "Resolvido Em"
            ]
            tickets_rows = []
            for t in tickets_db:
                tickets_rows.append([
                    t.id,
                    t.titulo,
                    t.cliente.nome_razao_social if t.cliente else "",
                    t.contrato_id or "",
                    t.status,
                    t.prioridade,
                    t.categoria,
                    t.descricao,
                    t.resolucao or "",
                    t.criado_por.full_name if t.criado_por else "",
                    t.atribuido_para.full_name if t.atribuido_para else "",
                    t.created_at,
                    t.resolvido_em
                ])
            zip_file.writestr("tickets_suporte.xlsx", cls._create_excel_file("Tickets de Suporte", tickets_headers, tickets_rows))

            # --- 7. Roteadores ---
            routers_db = db.query(Router).filter(Router.empresa_id == empresa_id).all()
            routers_headers = ["ID", "Nome", "IP", "Usuário", "Tipo", "Porta", "Ativo", "Criado Em"]
            routers_rows = []
            for r in routers_db:
                routers_rows.append([
                    r.id,
                    r.nome,
                    r.ip,
                    r.usuario,
                    r.tipo,
                    r.porta,
                    "Sim" if r.is_active else "Não",
                    r.created_at
                ])
            zip_file.writestr("roteadores.xlsx", cls._create_excel_file("Roteadores", routers_headers, routers_rows))

            # --- 8. Caixa (Sessões e Movimentações) ---
            sessoes_db = db.query(CaixaSessao).filter(CaixaSessao.empresa_id == empresa_id).all()
            sessao_ids = [s.id for s in sessoes_db]
            
            # Movimentações associadas às sessões da empresa
            movs_db = []
            if sessao_ids:
                movs_db = db.query(CaixaMovimentacao).filter(CaixaMovimentacao.sessao_id.in_(sessao_ids)).all()

            # Planilha 8a: Sessões de Caixa
            caixa_sessoes_headers = [
                "ID Sessão", "Operador", "Ponto de Venda", "Data Abertura", "Data Fechamento", 
                "Saldo Inicial", "Saldo Final Informado", "Saldo Final Calculado", "Status"
            ]
            caixa_sessoes_rows = []
            for s in sessoes_db:
                caixa_sessoes_rows.append([
                    s.id,
                    s.usuario.full_name if s.usuario else "",
                    s.local_pagamento.nome if s.local_pagamento else "",
                    s.data_abertura,
                    s.data_fechamento,
                    s.saldo_inicial,
                    s.saldo_final_informado or 0.0,
                    s.saldo_final_calculado or 0.0,
                    s.status
                ])
            zip_file.writestr("caixa_sessoes.xlsx", cls._create_excel_file("Sessões de Caixa", caixa_sessoes_headers, caixa_sessoes_rows))

            # Planilha 8b: Movimentações de Caixa
            caixa_movs_headers = [
                "ID Movimentação", "ID Sessão", "Operador", "Forma Pagamento", 
                "Tipo Movimentação", "Valor", "Descrição", "Data/Hora"
            ]
            caixa_movs_rows = []
            for m in movs_db:
                caixa_movs_rows.append([
                    m.id,
                    m.sessao_id,
                    m.usuario.full_name if m.usuario else "",
                    m.forma_pagamento.nome if m.forma_pagamento else "",
                    m.tipo,
                    m.valor,
                    m.descricao or "",
                    m.created_at
                ])
            zip_file.writestr("caixa_movimentacoes.xlsx", cls._create_excel_file("Movimentações de Caixa", caixa_movs_headers, caixa_movs_rows))

        zip_buffer.seek(0)
        return zip_buffer.getvalue()
