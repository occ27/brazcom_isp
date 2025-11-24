#!/usr/bin/env python3
# Simple converter for the XLSX cClass table to JSON
import json
import re
from openpyxl import load_workbook
from pathlib import Path

xlsx_path = Path(__file__).resolve().parents[1] / 'docs' / 'Tabela-cClass.xlsx'
if not xlsx_path.exists():
    print('XLSX not found:', xlsx_path)
    raise SystemExit(1)

wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
ws = wb.active

rows = list(ws.iter_rows(values_only=True))
if not rows:
    print('No rows found')
    raise SystemExit(1)

# Try to detect header row
header = [str(c).strip() if c is not None else '' for c in rows[0]]
# Common header names to look for
possible_code_headers = ['Grupo/Código', 'Código', 'Grupo/Codigo', 'Code', 'Grupo', 'Codigo']
possible_title_headers = ['Descrição Principal', 'Descrição', 'Descricao', 'Title', 'Descrição Principal']
possible_example_headers = ['Exemplo de Subitem', 'Exemplo', 'Example']

code_idx = None
title_idx = None
example_idx = None
for i, h in enumerate(header):
    if any(k.lower() == h.lower() for k in possible_code_headers):
        code_idx = i
    if any(k.lower() == h.lower() for k in possible_title_headers):
        title_idx = i
    if any(k.lower() == h.lower() for k in possible_example_headers):
        example_idx = i

# Fallback: guess positions if not found
if code_idx is None:
    code_idx = 0
if title_idx is None:
    title_idx = 1 if len(header) > 1 else 0

items = []
for r in rows[1:]:
    if all(cell is None for cell in r):
        continue
    raw_code = r[code_idx] if code_idx < len(r) else None
    raw_title = r[title_idx] if title_idx < len(r) else None
    raw_example = r[example_idx] if example_idx is not None and example_idx < len(r) else None
    if raw_code is None and raw_title is None:
        continue
    code = str(raw_code).strip()
    # normalize code to preserve leading zeros (keep as string)
    # remove decimal part if Excel interpreted as number like 10.0
    if re.match(r'^\d+\.0$', code):
        code = code.split('.')[0]
    # remove any non-digit characters
    code = re.sub(r'[^0-9]', '', code)
    # pad to 3 or 4 digits? We'll keep as-is but preserve leading zeros if present
    title = str(raw_title).strip() if raw_title is not None else ''
    example = str(raw_example).strip() if raw_example is not None else ''
    # sanitize weird characters that sometimes come from Excel conversion
    title = title.replace('\u2019', "'").replace('\u201c', '"').replace('\u201d', '"')
    items.append({
        'code': code.zfill(len(code)) if code else '',
        'title': title,
        'example': example,
    })

out_path = Path(__file__).resolve().parents[1] / 'frontend' / 'src' / 'data' / 'cclass.json'
out_path.parent.mkdir(parents=True, exist_ok=True)
with open(out_path, 'w', encoding='utf-8') as f:
    json.dump(items, f, ensure_ascii=False, indent=2)

print('Wrote', out_path)
